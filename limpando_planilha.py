import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Color
import PySimpleGUI as sg
import os

class limpar:
    def __init__(self):
        pass
    def programa(self, arquivo):
        nome = 'Resultado Clientes'
        df = pd.read_excel(arquivo)
        # Formatação dos CNPJs
        nome_da_coluna = 'CNPJ'
        print(df.columns)
        if nome_da_coluna in df.columns:
            numero_da_coluna = df.columns.get_loc(nome_da_coluna)
            for indice, valor in enumerate(df['CNPJ']):
                if pd.notnull(valor) and isinstance(valor, int):
                    valor = str(valor)
                    if len(valor) == 14:
                        cnpj_formatado = f"{valor[:2]}.{valor[2:5]}.{valor[5:8]}/{valor[8:12]}-{valor[12:]}"
                    else:
                        valor = "0" + valor
                        cnpj_formatado = f"{valor[:2]}.{valor[2:5]}.{valor[5:8]}/{valor[8:12]}-{valor[12:]}"
                    df.iloc[indice, numero_da_coluna] = cnpj_formatado
                if isinstance(valor, str):
                    pass
                else:
                    df.iloc[indice, numero_da_coluna] = ""
                

        # Salvar DataFrame atualizado no Excel
        df.to_excel(f'{nome}.xlsx', index=False)

        # Carregar o arquivo Excel tratado
        wb = load_workbook(f'{nome}.xlsx')
        sheet = wb.active

        font_bold = Font(bold=True)  # Negrito
        # Define a cor de preenchimento (azul marinho)
        fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        # Defina um estilo de borda dupla e azul em todos os lados
        border_style = Border(
            top=Side(style='double', color=Color(rgb='4F81BD')),  # Borda dupla azul no topo
            bottom=Side(style='double', color=Color(rgb='4F81BD'))  # Borda dupla azul na parte inferior
        )

        # Formatar a primeira linha (índice) em negrito e com a cor de preenchimento
        for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.font = font_bold
                cell.fill = fill
                

        # Formatar as linhas com "CONSOLIDADO" em negrito
        for indice, valor in enumerate(df['Empresa'], start=2):  # Começando a partir da segunda linha (linha 1 é o cabeçalho)
            # Verificar se o valor contém a substring desejada
            substring_procurada = "CONSOLIDADO"  # Nome a ser procurado
            
            # Se o valor contém a substring "CONSOLIDADO", aplicar formatação em negrito à linha
            if substring_procurada in str(valor).upper():
                for col_num, col_name in enumerate(df.columns, start=1):
                    cell = sheet.cell(row=indice, column=col_num)
                    cell.font = font_bold
                    cell.fill = fill
                
        # Função para pintar números negativos de vermelho e aplicar negrito
        def format_negative_value(cell):
            if cell.value is not None and isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.font = Font(bold=True, color="FF0000")  # Negrito e cor vermelha

        # Aplicar formatação às células com números negativos
        for row in sheet.iter_rows(min_row=2, min_col=5, max_col=sheet.max_column):
            for cell in row:
                format_negative_value(cell)

        # Percorra todas as linhas e colunas da planilha
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border_style

        # Salvar o arquivo Excel com as linhas em negrito
        wb.save(f'{nome}.xlsx')

if __name__ == "__main__":
    tratamento = limpar()
    selected_theme = 'Reddit'
    sg.theme(selected_theme)
    file = sg.popup_get_file('Selecione o arquivo bruto',  title="Resultado Clientes - Nova Versão", keep_on_top=True,icon=os.path.abspath('iconp.ico'))
    if file == None:
        sg.popup_error('Arquivo não selecionado!', title=" ", keep_on_top=True,icon=os.path.abspath('iconp.ico'))
    else:
        tratamento.programa(file)
        