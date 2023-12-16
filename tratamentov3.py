import pandas as pd
from openpyxl import load_workbook
import datetime
import os
import locale
import time
from openpyxl.styles import PatternFill

# Iniciar o temporizador
start_time = time.time()

# Configurar a localização para o Brasil (pt_BR)
locale.setlocale(locale.LC_ALL, 'pt_BR')

# Obter a data atual
data_atual = datetime.date.today()
ano = data_atual.year

'''inicio da tratativa do dataframe
--------------------------------------------------------------------------------------------'''
# Lê o arquivo 'dre.xlsx' e armazena os dados em um DataFrame df.
#df = pd.read_excel(os.path.abspath('dre.xlsx'))
df = pd.read_excel(os.path.abspath('dre.xlsx'))

movimentacao = df.iloc[3,0]
coluna_movimentacao = movimentacao.split(' / ')
coluna_movimentacao = coluna_movimentacao[1].split(': ')[1]

unidade = df.iloc[4,0]
coluna_unidade = unidade.split(' / ')
coluna_unidade = coluna_unidade[0]

# Remover as primeiras 8 linhas
df = df.drop(range(0, 8))
# Redefinir o índice
df = df.reset_index(drop=True)

# Concatenar as duas primeiras linhas para formar o novo cabeçalho
primeira_linha = df.iloc[0].astype(str)
segunda_linha = df.iloc[1].astype(str)
novo_cabecalho = primeira_linha + '-' + segunda_linha
#define a primeira linha como cabeçado e remove os espaço dela
novo_cabecalho = novo_cabecalho.str.replace(' ', '').str.replace('-', ' ').str.replace('nan', '').str.replace('ç', 'c').str.replace('ã', 'a')
df.columns = novo_cabecalho

'''--Lista de meses--'''
# Extrair apenas os meses usando uma expressão regular
segunda_linha_meses = segunda_linha.str.extract(r'(\w{3})/\d{4}').squeeze()
# Remover valores duplicados e valores 'NaN' da série e transformar em uma lista
meses = segunda_linha_meses.drop_duplicates().dropna().tolist()

# Filtrar colunas que contêm a substring 'Saldo', 'Débitos', 'Créditos', 'Metas/Orçam.' e '%Mt/Or'
colunas_a_remover = df.columns[df.columns.str.contains('Saldo|Débitos|Créditos|Metas/Orcam.|%Mt/Or')]
df = df.drop(columns=colunas_a_remover)

# Redefinir o índice
df = df.reset_index(drop=True)
df = df.drop(range(0, 2))

#procurando valores CR,DB e removendo de cada celula
def clean_value(x):
    if isinstance(x, str) and 'DB' in x:
        x = x.replace('DB', '').strip().replace(".", "").replace(",", ".")
        x = "-" + x
        #x = float(x) * -1
        return x
    elif isinstance(x, str) and 'CR' in x:
        x = x.replace('CR', '').strip().replace(".", "").replace(",", ".")
        #x = float(x)
        return x
    return x

df.iloc[:, 4:] = df.iloc[:, 4:].applymap(clean_value)

'''Inicio da procura e calculo necessario
            --------------------------------------------------------------------------------------------'''
#dicionario a ser preenchido
#Planilha DRE
contas_resultado = {}
provisao_de_imposto = {}
receitas_operacaional = {}
receitas_operacaional = {}
venda_mercadorias = {}
recuperacao_despesas = {}
contribuicao_social = {}
recuperacao_despesa = {}
quebra_estoque = {} #
outras_despesas_vendas = {} #
ajuste_inventario = {}
despesas_operacional = {} #
receita_operacional = {} #
custo_mercadorias = {} #
receitas_comercial = {}  #
perdas_mercadorias = {} #
despesas_embalagem = {} #
depreciacoes = {}
juros_financiamento = {}
juros_capital_giro = {}
resultado_exercicio = {}
receitas_diversas = {}


# Resumo    
# resultado_liquido = {}
# margem_contabil = {}
# lucro = {}
# receitas_nao_operacionais = {}
# recolhido = {}
# lucro_contabil = {}
# venda = {}
# ir = {}
# add = {}
# csll = {}
# total = {}

#resumo v3
receita_operacionais = {} 
custo_mercadorias_vendido = {}
margem_bruta = {}
receitas_comerciais = {}  
margem_comerciais = {}
quebra_cotabil = {}
despesas_venda = {}
margem_operacional = {}
despesas_operacionais = {} 
ebitda = {}
depreciacao = {}
juros_financeiros = {}
receitas_financeiras = {}
lair = {}
ir = {}
csll = {}    
lucro_liquido = {}

#porcentagem v3
receita_operacionais_porc = {} 
custo_mercadorias_vendido_porc = {}
margem_bruta_porc = {}
receitas_comerciais_porc = {}  
margem_comerciais_porc = {}
quebra_cotabil_porc = {}
despesas_venda_porc = {}
margem_operacional_porc = {}
despesas_operacionais_porc = {} 
ebitda_porc = {}
depreciacao_porc = {}
juros_financeiros_porc = {}
receitas_financeiras_porc = {}
lair_porc = {}
ir_porc = {}
csll_porc = {}    
lucro_liquido_porc = {}

#trimestre v3
primeiro_trimestre = {} 
segundo_trimestre = {}
terceiro_trimestre = {}
quarto_trimestre = {}

primeiro_trimestre_porc = {} 
segundo_trimestre_porc = {}
terceiro_trimestre_porc = {}
quarto_trimestre_porc = {}

def valor_da_linha(linha_descricao, coluna_a_encontrar,contem=True):
    coluna_filtro = 'DescricaoConta '
    try:
        # Tratar valores ausentes substituindo por uma string vazia
        df[coluna_filtro] = df[coluna_filtro].fillna('')
        if contem:
            #encontrar linha a partir de uma coluna como filtro e contendo os valores
            linha_contas_resultado = df[df[coluna_filtro].str.contains(linha_descricao, case=False)]
        else:
            # Localizar a célula que contém o texto inserido
            linha_contas_resultado = df[df[coluna_filtro].str.strip() == linha_descricao]
        # Obter o valor da coluna inserido
        valor_linha = linha_contas_resultado[coluna_a_encontrar].values[0]
        indice_linha = linha_contas_resultado.index[linha_contas_resultado[coluna_a_encontrar] == valor_linha].tolist()[0]
        num_coluna = df.columns.get_loc(coluna_a_encontrar)
        valor_linha = linha_contas_resultado[coluna_a_encontrar].str.strip().values[0]
        if valor_linha == '':
            valor_linha = 0
        else:
            valor_linha = float(linha_contas_resultado[coluna_a_encontrar].values[0])
        return valor_linha, indice_linha , num_coluna
    except Exception as e:
        return 0
    
for mes in meses:
    #definindo o ano e mes para procurar em relação a função mês
    coluna_a_encontrar = f'MvtoLíquido {mes}/{ano}'
    
    def preencher_dicionario(linha, dicionario, contem_linha=True):
        valor_linha, indice_linha, num_coluna = valor_da_linha(linha, coluna_a_encontrar, contem_linha)
        dicionario[f'MvtoLíquido {mes}/{ano}'] = valor_linha
        dicionario[f'Index_linha {mes}/{ano}'] = indice_linha
        dicionario[f'Index_coluna {mes}/{ano}'] = num_coluna
        
    #atualizando o dicionario
    def att(dataset,resultado):
        for item in dataset.items():
            indice1_dicionario = item[:2]
            if indice1_dicionario[1] == dataset[coluna_a_encontrar]: 
                dataset.update({ item[0]: resultado })
                break

    preencher_dicionario('CUSTO DAS MERCADORIAS VENDIDOS', custo_mercadorias)          
    preencher_dicionario('CONTAS DE RESULTADO',contas_resultado)
    preencher_dicionario('PROVISAO DE IMPOSTO S/L', provisao_de_imposto)
    preencher_dicionario('RECEITAS OPERACIONAL LIQUIDA', receitas_operacaional)
    preencher_dicionario('VENDA DE MERCADORIAS',venda_mercadorias)
    preencher_dicionario('Recuperacao De Despesas Exerc Anterior',recuperacao_despesas)
    preencher_dicionario('Contribuicao Social',contribuicao_social,False)
    preencher_dicionario('Recuperacao De Despesas Exerc Anterior', recuperacao_despesa)
    preencher_dicionario('RECEITAS OPERACIONAL LIQUIDA',receita_operacional,False)
    preencher_dicionario('RECEITAS COMERCIAIS',receitas_comercial,False)
    preencher_dicionario('Quebras De Estoque',quebra_estoque)
    preencher_dicionario('DESPESAS OPERACIONAIS',despesas_operacional,False)
    preencher_dicionario('OUTRAS DESPESAS VENDAS',outras_despesas_vendas,False)
    preencher_dicionario('DESPESAS EMBALAGENS',despesas_embalagem,False)
    preencher_dicionario('Depreciacoes',depreciacoes,False)
    preencher_dicionario('RECEITAS DIVERSAS',receitas_diversas,False)
    preencher_dicionario('RESULTADO DO EXERCICIO',resultado_exercicio,False)
    try:
        preencher_dicionario('Juros S/ Capital de Giro',juros_financiamento,False)
        preencher_dicionario('Juros s/Financiamento FCO',juros_capital_giro,False)
        preencher_dicionario('Ajuste de Inventário',ajuste_inventario)
    except:
        preencher_dicionario('Perdas de Mercadorias',perdas_mercadorias)
        preencher_dicionario('Juros S/ Financiamento',juros_capital_giro,False)

    #contas de resultado calculo                 
    resultado_analisado = contas_resultado[coluna_a_encontrar] + provisao_de_imposto[coluna_a_encontrar] * - 1
    att(contas_resultado,resultado_analisado)


# Salvar planilha Excel tratada, trocando index = True mostra o index das linhas
df.to_excel('dre_tratada.xlsx', index=False)

'''definindo a primeira linha como cabeçalho, copiando o dataframe e criando novas aba com os nomes (Resumo, Planilha DRE)
            --------------------------------------------------------------------------------------------'''
# Carregar o arquivo Excel tratado
wb = load_workbook('dre_tratada.xlsx')
# Renomear a planilha 'Sheet1' para 'Planilha DRE'
if 'Sheet1' in wb.sheetnames:
    sheet = wb['Sheet1']
    sheet.title = 'Planilha DRE'
# Cria uma nova planilha antes da planilha ativa
wb.create_sheet('Resumo', 0)
# Salvar novamente o arquivo
wb.save('dre_tratada.xlsx')

# Lê o arquivo 'dre_tratada.xlsx' e armazena os dados em um DataFrame df_sheet1
df_sheet1 = pd.read_excel(os.path.abspath('dre_tratada.xlsx'), sheet_name='Planilha DRE')
# Lê o arquivo 'dre_tratada.xlsx' e armazena os dados em um DataFrame df_resumo
df_resumo = pd.read_excel(os.path.abspath('dre_tratada.xlsx'), sheet_name='Resumo')
df_resumo.fillna(method='ffill', inplace=True)

# Redefinir o índice
df_sheet1 = df_sheet1.reset_index(drop=True)

# Definindo os cabeçalhos resumo
cabecalhos = pd.Series(segunda_linha[segunda_linha != 'nan'].str.strip().drop_duplicates())
cabecalhos['Unnamed: 2'] = 'DescriçãoConta'

# Mapeamento de valores 'value' para novas chaves e trimestres
mapeamento_trimestres = {
    f'Mar/{ano}': ('Unnamed: 666', f'1º Trimestre {ano}'),
    f'Jun/{ano}': ('Unnamed: 333', f'2º Trimestre {ano}'),
    f'Set/{ano}': ('Unnamed: 166', f'3º Trimestre {ano}'),
    f'Dez/{ano}': ('Unnamed: 83', f'4º Trimestre {ano}')
}
#acrescentando trimestre no indice
for col, value in cabecalhos.items():
    if value in mapeamento_trimestres:
        nova_chave, trimestre = mapeamento_trimestres[value]
        indice_marco = list(cabecalhos.keys()).index(col)
        cabecalhos = dict(list(cabecalhos.items())[:indice_marco+1] + [(nova_chave, trimestre)] + list(cabecalhos.items())[indice_marco:])

mapeamento_meses = {
    f'Jan/{ano}': ('Unnamed: 50', 'Jan %'),
    f'Fev/{ano}': ('Unnamed: 51', 'Fev %'),
    f'Mar/{ano}': ('Unnamed: 52', 'Mar %'),
    f'1º Trimestre {ano}': ('Unnamed: 53', '1º Tri %'),
    f'Abr/{ano}': ('Unnamed: 54', 'Abr %'),
    f'Mai/{ano}': ('Unnamed: 55', 'Mai %'),
    f'Jun/{ano}': ('Unnamed: 56', 'Jun %'),
    f'2º Trimestre {ano}': ('Unnamed: 57', '2º Tri %'),
    f'Jul/{ano}': ('Unnamed: 58', 'Jul %'),
    f'Ago/{ano}': ('Unnamed: 59', 'Ago %'),
    f'Set/{ano}': ('Unnamed: 60', 'Set %'),
    f'3º Trimestre {ano}': ('Unnamed: 61', '3º Tri %'),
    f'Out/{ano}': ('Unnamed: 62', 'Out %'),
    f'Nov/{ano}': ('Unnamed: 63', 'Nov %'),
    f'Dez/{ano}': ('Unnamed: 64', 'Dez %'),
    f'4º Trimestre {ano}': ('Unnamed: 65', '4º Tri %')
    }

#acrescentando porcentagem em cada mes no indice
for col, value in cabecalhos.items():
    if value in mapeamento_meses:
        nova_chave, porcentagem = mapeamento_meses[value]
        indice_marco = list(cabecalhos.keys()).index(col)
        cabecalhos = dict(list(cabecalhos.items())[:indice_marco+1] + [(nova_chave, porcentagem)] + list(cabecalhos.items())[indice_marco:])

# Preenchendo os cabeçalhos na primeira linha do DataFrame
for col, cabecalho in enumerate(cabecalhos.values()):
    df_resumo.at[1, col] = cabecalho

# Adicionando a coluna vazia antes dos cabeçalhos
df_resumo.insert(0, '', '')

linhas = ['RECEITA OPERACIONAL LÍQUIDA','Custo de Mercadoria Vendida','MARGEM BRUTA','Receitas Comerciais','MARGEM COMERCIAL','Quebra Contábil',
          'Despesa C/ Venda','MARGEM OPERACIONAL','Despesas Operacionais','EBITDA','Depreciação','Juros Financeiros','Receita Financeiras',
          'LAIR','IR','CSLL','LL - Lucro Líquido']

for lin, linha in enumerate(linhas):
    df_resumo.at[lin + 2, 0] = linha

# Definir a primeira linha como cabeçalho
novo_cabecalho = df_resumo.iloc[0].astype(str)
df_resumo.columns = novo_cabecalho
# Remover a primeira linha, que agora é o cabeçalho duplicado
df_resumo = df_resumo.iloc[1:].reset_index(drop=True)

def valor_da_linha_resumo(linha_descricao, coluna_a_encontrar,contem=True,porcentagem=False):
    coluna_filtro = 'DescriçãoConta'
    try:
        # Tratar valores ausentes substituindo por uma string vazia
        df_resumo[coluna_filtro] = df_resumo[coluna_filtro].fillna('')
        if contem:
            #encontrar linha a partir de uma coluna como filtro e contendo os valores
            linha_contas_resultado = df_resumo[df_resumo[coluna_filtro].str.contains(linha_descricao, case=False)]
        else:
            # Localizar a célula que contém o texto inserido
            linha_contas_resultado = df_resumo[df_resumo[coluna_filtro].str.strip() == linha_descricao]
        
        if porcentagem:
            valor_linha = linha_contas_resultado[coluna_a_encontrar].values[0]
            indice_linha = linha_contas_resultado.index[0]
            num_coluna = df_resumo.columns.get_loc(coluna_a_encontrar) + 1
        else:
            # Obter o valor da coluna inserido
            valor_linha = linha_contas_resultado[coluna_a_encontrar].values[0]
            indice_linha = linha_contas_resultado.index[0]
            try:
                num_coluna = df_resumo.columns.get_loc(coluna_a_encontrar)
            except:
                if coluna_a_encontrar[0] == f'1º Trimestre {ano}':
                    num_coluna = 8
                elif coluna_a_encontrar[0] == f'2º Trimestre {ano}':
                    num_coluna = 16
                elif coluna_a_encontrar[0] == f'3º Trimestre {ano}':
                    num_coluna = 14
                elif coluna_a_encontrar[0] == f'4º Trimestre {ano}':
                    num_coluna = 32
                elif coluna_a_encontrar[0] == '1º Tri %':
                    num_coluna = 9
                elif coluna_a_encontrar[0] == '2º Tri %':
                    num_coluna = 17
                elif coluna_a_encontrar[0] == '3º Tri %':
                    num_coluna = 15
                elif coluna_a_encontrar[0] == '4º Tri %':
                    num_coluna = 33
                
        if type(valor_linha) == str:
            valor_linha = 0
        return valor_linha, indice_linha , num_coluna
    except Exception as e:
        return 0
    
for mes in meses:
    coluna_a_encontrar = f'{mes}/{ano}'
    
    def preencher_dicionario_resumo(linha, dicionario, contem_linha=True):
        venda, index_linha, index_coluna = valor_da_linha_resumo(linha, coluna_a_encontrar,contem_linha)
        dicionario[f'MvtoLíquido {mes}/{ano}'] = venda
        dicionario[f'Index_linha {mes}/{ano}'] = index_linha
        dicionario[f'Index_coluna {mes}/{ano}'] = index_coluna
    
    def preencher_dicionario_resumo_porc(linha, dicionario, contem_linha=True,porc=True):
        venda, index_linha, index_coluna = valor_da_linha_resumo(linha, coluna_a_encontrar,contem_linha,porc)
        dicionario[f'{mes} %'] = venda
        dicionario[f'Index_linha {mes}/{ano}'] = index_linha
        dicionario[f'Index_coluna {mes}/{ano}'] = index_coluna

    def preencher_valor_planilha(linha,calculo=None):
            try:
                linha_ = int(linha[f'Index_linha {mes}/{ano}'])
                coluna_ = int(linha[f'Index_coluna {mes}/{ano}'])
                df_resumo.iloc[linha_,coluna_] = calculo
            except:
               
                try:
                    elementos_para_atualizar = {
                        'RECEITA OPERACIONAL LÍQUIDA': 'RECEITA OPERACIONAL LÍQUIDA',
                        'Custo de Mercadoria Vendida': 'Custo de Mercadoria Vendida',
                        'MARGEM BRUTA': 'MARGEM BRUTA',
                        'Receitas Comerciais': 'Receitas Comerciais',
                        'MARGEM COMERCIAL': 'MARGEM COMERCIAL',
                        'Quebra Contábil': 'Quebra Contábil',
                        'Despesa C/ Venda': 'Despesa C/ Venda',
                        'MARGEM OPERACIONAL': 'MARGEM OPERACIONAL',
                        'Despesas Operacionais': 'Despesas Operacionais',
                        'EBITDA': 'EBITDA',
                        'Depreciação': 'Depreciação',
                        'Juros Financeiros': 'Juros Financeiros', 
                        'Receita Financeiras': 'Receita Financeiras',
                        'LAIR': 'LAIR',
                        'IR': 'IR', 
                        'CSLL': 'CSLL',
                        'LL - Lucro Líquido': 'LL - Lucro Líquido'
                    }

                    for elemento, chave in elementos_para_atualizar.items():
                        if elemento in primeiro_trimestre:
                            linha_ = int(linha[f'{chave} linha {ano}'])
                            coluna_ = int(linha[f'{chave} coluna {ano}'])
                            if not pd.isna(linha[f'{elemento}']): 
                            # Converta o valor não nulo para inteiro
                                df_resumo.iloc[linha_, coluna_] = float(linha[f'{elemento}'])
                except:
                    pass
                
    preencher_dicionario_resumo('RECEITA OPERACIONAL LÍQUIDA',receita_operacionais,False)
    preencher_dicionario_resumo('Custo de Mercadoria Vendida',custo_mercadorias_vendido,False)
    preencher_dicionario_resumo('MARGEM BRUTA',margem_bruta,False)
    preencher_dicionario_resumo('Receitas Comerciais',receitas_comerciais,False)
    preencher_dicionario_resumo('MARGEM COMERCIAL',margem_comerciais,False)
    preencher_dicionario_resumo('Quebra Contábil',quebra_cotabil,False)
    preencher_dicionario_resumo('Despesa C/ Venda',despesas_venda,False)
    preencher_dicionario_resumo('MARGEM OPERACIONAL',margem_operacional,False)
    preencher_dicionario_resumo('Despesas Operacionais',despesas_operacionais,False)
    preencher_dicionario_resumo('EBITDA',ebitda,False)
    preencher_dicionario_resumo('Depreciação',depreciacao,False)
    preencher_dicionario_resumo('Juros Financeiros',juros_financeiros,False)
    preencher_dicionario_resumo('Receita Financeiras',receitas_financeiras,False)
    preencher_dicionario_resumo('LAIR',lair,False)
    preencher_dicionario_resumo('IR',ir,False)
    preencher_dicionario_resumo('CSLL',csll,False)
    preencher_dicionario_resumo('LL - Lucro Líquido',lucro_liquido,False)

    preencher_dicionario_resumo_porc('RECEITA OPERACIONAL LÍQUIDA',receita_operacionais_porc,False)
    preencher_dicionario_resumo_porc('Custo de Mercadoria Vendida',custo_mercadorias_vendido_porc,False)
    preencher_dicionario_resumo_porc('MARGEM BRUTA',margem_bruta_porc,False)
    preencher_dicionario_resumo_porc('Receitas Comerciais',receitas_comerciais_porc,False)
    preencher_dicionario_resumo_porc('MARGEM COMERCIAL',margem_comerciais_porc,False)
    preencher_dicionario_resumo_porc('Quebra Contábil',quebra_cotabil_porc,False)
    preencher_dicionario_resumo_porc('Despesa C/ Venda',despesas_venda_porc,False)
    preencher_dicionario_resumo_porc('MARGEM OPERACIONAL',margem_operacional_porc,False)
    preencher_dicionario_resumo_porc('Despesas Operacionais',despesas_operacionais_porc,False)
    preencher_dicionario_resumo_porc('EBITDA',ebitda_porc,False)
    preencher_dicionario_resumo_porc('Depreciação',depreciacao_porc,False)
    preencher_dicionario_resumo_porc('Juros Financeiros',juros_financeiros_porc,False)
    preencher_dicionario_resumo_porc('Receita Financeiras',receitas_financeiras_porc,False)
    preencher_dicionario_resumo_porc('LAIR',lair_porc,False)
    preencher_dicionario_resumo_porc('IR',ir_porc,False)
    preencher_dicionario_resumo_porc('CSLL',csll_porc,False)
    preencher_dicionario_resumo_porc('LL - Lucro Líquido',lucro_liquido_porc,False)
    
    
    #RECEITA OPERACIONAL LÍQUIDA
    preencher_valor_planilha(receita_operacionais,receita_operacional[f'MvtoLíquido {mes}/{ano}'])
    calculo_receita_porc = receita_operacional[f'MvtoLíquido {mes}/{ano}'] / receita_operacional[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(receita_operacionais_porc,calculo_receita_porc)

    #Custo de Mercadoria Vendida
    preencher_valor_planilha(custo_mercadorias_vendido,custo_mercadorias[f'MvtoLíquido {mes}/{ano}'])
    calculo_mercadoria_porc = custo_mercadorias[f'MvtoLíquido {mes}/{ano}'] / receita_operacional[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(custo_mercadorias_vendido_porc,calculo_mercadoria_porc)

    #MARGEM BRUTA
    calculo_margem_bruta = receita_operacional[f'MvtoLíquido {mes}/{ano}'] + custo_mercadorias[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(margem_bruta,calculo_margem_bruta)
    calculo_margem_bruta_porc = calculo_margem_bruta / receita_operacional[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(margem_bruta_porc,calculo_margem_bruta_porc)

    #Receitas Comerciais
    preencher_valor_planilha(receitas_comerciais,receitas_comercial[f'MvtoLíquido {mes}/{ano}'])
    calculo_receita_comercial_porc = receitas_comercial[f'MvtoLíquido {mes}/{ano}'] / receita_operacional[f'MvtoLíquido {mes}/{ano}'] 
    preencher_valor_planilha(receitas_comerciais_porc,calculo_receita_comercial_porc)

    #MARGEM COMERCIAL
    calculo_margem = calculo_margem_bruta + receitas_comercial[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(margem_comerciais,calculo_margem)
    calculo_margem_porc = calculo_margem / receita_operacional[f'MvtoLíquido {mes}/{ano}'] 
    preencher_valor_planilha(margem_comerciais_porc,calculo_margem_porc)

    #Quebra Contábil - revisar celula D259 - Ajuste de Inventário
    try:
        calculo_quebra_cotabil = quebra_estoque[f'MvtoLíquido {mes}/{ano}'] + ajuste_inventario[f'MvtoLíquido {mes}/{ano}']
    except:
        calculo_quebra_cotabil = quebra_estoque[f'MvtoLíquido {mes}/{ano}'] + perdas_mercadorias[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(quebra_cotabil,calculo_quebra_cotabil)
    calculo_quebra_contabil_porc = calculo_quebra_cotabil / receita_operacional[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(quebra_cotabil_porc,calculo_quebra_contabil_porc)
    #Despesas c/ venda
    try:
        calculo_despesas = despesas_embalagem[f'MvtoLíquido {mes}/{ano}']
    except:
        calculo_despesas =  0
    preencher_valor_planilha(despesas_venda,despesas_embalagem[f'MvtoLíquido {mes}/{ano}'])
    calculo_despesas_porc = calculo_despesas / receita_operacional[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(despesas_venda_porc,calculo_despesas_porc)
    #MARGEM OPERACIONAL 
    calculo_margem_operacional = calculo_margem  + calculo_quebra_cotabil + calculo_despesas
    preencher_valor_planilha(margem_operacional,calculo_margem_operacional)
    calculo_margem_operacional_porc = calculo_margem_operacional / receita_operacional[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(margem_operacional_porc,calculo_margem_operacional_porc)
    #Despesas Operacionais
    #calculo em base de -(DESPESAS OPERACIONAIS + Quebra Contábil + Despesa C/ Vend + Juros Financeiros + Depreciação)
    
    try:
        calculo_juros_financeiro = juros_financiamento[f'MvtoLíquido {mes}/{ano}'] + juros_capital_giro[f'MvtoLíquido {mes}/{ano}']
    except:
        calculo_juros_financeiro = juros_capital_giro[f'MvtoLíquido {mes}/{ano}']

    calculo_despesas_operacional = despesas_operacional[f'MvtoLíquido {mes}/{ano}'] + calculo_quebra_cotabil + calculo_despesas + depreciacoes[f'MvtoLíquido {mes}/{ano}'] + calculo_juros_financeiro
    preencher_valor_planilha(despesas_operacionais,calculo_despesas_operacional)
    calculo_despesas_operacional_porc = calculo_despesas_operacional / receita_operacional[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(despesas_operacionais_porc,calculo_despesas_operacional_porc)

    #Ebitda
    calculo_ebitida = calculo_despesas_operacional + calculo_margem_operacional
    preencher_valor_planilha(ebitda,calculo_ebitida)
    calculo_ebitida_porc = calculo_ebitida / receita_operacional[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(ebitda_porc,calculo_ebitida_porc)
    #depeciação
    preencher_valor_planilha(depreciacao,depreciacoes[f'MvtoLíquido {mes}/{ano}'])
    calculo_depreciacao_porc = depreciacoes[f'MvtoLíquido {mes}/{ano}'] / receita_operacional[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(depreciacao_porc,calculo_depreciacao_porc)
    #juros financeiros
    preencher_valor_planilha(juros_financeiros,calculo_juros_financeiro)
    calculo_juros_financeiro_porc = calculo_juros_financeiro / receita_operacional[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(juros_financeiros_porc,calculo_juros_financeiro_porc)
    #receitas financeiras
    calculo_financeiras = resultado_exercicio[f'MvtoLíquido {mes}/{ano}'] + receitas_diversas[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(receitas_financeiras,calculo_financeiras)
    calculo_financeiras_porc = calculo_financeiras / receita_operacional[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(receitas_financeiras_porc,calculo_financeiras_porc)
    #LAIR
    calculo_lair = calculo_ebitida + depreciacoes[f'MvtoLíquido {mes}/{ano}'] + calculo_juros_financeiro + calculo_financeiras
    preencher_valor_planilha(lair,calculo_lair)
    calculo_lair_porc = calculo_lair / receita_operacional[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(lair_porc,calculo_lair_porc)
    #IR
    calculo_IR = 0
    preencher_valor_planilha(ir,calculo_IR)
    calculo_IR_porc = calculo_IR / receita_operacional[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(ir_porc,calculo_IR_porc)

    #CSLL
    calculo_CSSL = calculo_lair * 0.09
    preencher_valor_planilha(csll,calculo_CSSL)
    calculo_CSSL_porc = calculo_CSSL / receita_operacional[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(csll_porc,calculo_CSSL_porc)
    #LL - Lucro Líquido 
    calculo_lucro_liquido = calculo_lair + calculo_IR + calculo_CSSL
    preencher_valor_planilha(lucro_liquido,calculo_lucro_liquido)
    calculo_lucro_liquido_porc = calculo_lucro_liquido / receita_operacional[f'MvtoLíquido {mes}/{ano}']
    preencher_valor_planilha(lucro_liquido_porc,calculo_lucro_liquido_porc)

###########################################################
for i in range(1,4):

    coluna_a_encontrar_tri = [f'{i}º Trimestre {ano}']
    coluna_a_encontrar_tri_porc = [f'{i}º Tri %'] 

    mapeamento_linha_chave = {
        'RECEITA OPERACIONAL LÍQUIDA': 'RECEITA OPERACIONAL LÍQUIDA',
        'Custo de Mercadoria Vendida': 'Custo de Mercadoria Vendida',
        'MARGEM BRUTA': 'MARGEM BRUTA',
        'Receitas Comerciais': 'Receitas Comerciais',
        'MARGEM COMERCIAL': 'MARGEM COMERCIAL',
        'Quebra Contábil': 'Quebra Contábil',
        'Despesa C/ Venda': 'Despesa C/ Venda',
        'MARGEM OPERACIONAL': 'MARGEM OPERACIONAL',
        'Despesas Operacionais': 'Despesas Operacionais',
        'EBITDA': 'EBITDA',
        'Depreciação': 'Depreciação',
        'Juros Financeiros': 'Juros Financeiros',
        'Receita Financeiras': 'Receita Financeiras',
        'LAIR': 'LAIR',
        'IR': 'IR',
        'CSLL': 'CSLL',
        'LL - Lucro Líquido': 'LL - Lucro Líquido'
    }

    def preencher_dicionarios(dicionario, dicionario_porc, colunas, colunas_porc, contem_linha=True):
        linhas = [
            'RECEITA OPERACIONAL LÍQUIDA', 'Custo de Mercadoria Vendida', 'MARGEM BRUTA',
            'Receitas Comerciais', 'MARGEM COMERCIAL', 'Quebra Contábil',
            'Despesa C/ Venda', 'MARGEM OPERACIONAL', 'Despesas Operacionais',
            'EBITDA', 'Depreciação', 'Juros Financeiros', 'Receita Financeiras',
            'LAIR', 'IR', 'CSLL', 'LL - Lucro Líquido'
        ]

        for linha in linhas:
            try:
                venda, index_linha, index_coluna = valor_da_linha_resumo(linha, colunas, contem_linha)
                chave = mapeamento_linha_chave.get(linha)
                if chave:
                    dicionario[chave] = venda
                    dicionario[f'{chave} linha {ano}'] = index_linha
                    dicionario[f'{chave} coluna {ano}'] = index_coluna
            except:
                pass
        for linha in linhas:
            try:
                venda, index_linha, index_coluna = valor_da_linha_resumo(linha, colunas_porc, contem_linha)
                chave = mapeamento_linha_chave.get(linha)
                if chave:
                    dicionario_porc[chave] = venda
                    dicionario_porc[f'{chave} linha {ano}'] = index_linha
                    dicionario_porc[f'{chave} coluna {ano}'] = index_coluna
            except:
                pass
    
    
    if i == 1:
        preencher_dicionarios(primeiro_trimestre, primeiro_trimestre_porc, coluna_a_encontrar_tri, coluna_a_encontrar_tri_porc, False)
        #RECEITA OPERACIONAL LÍQUIDA
        receita_trimestre = float(receita_operacional[f'MvtoLíquido Jan/{ano}'] + receita_operacional[f'MvtoLíquido Fev/{ano}'] + receita_operacional[f'MvtoLíquido Mar/{ano}'])
        receita_trimestre_porc = receita_trimestre / receita_trimestre

        #Custo de Mercadoria Vendida
        custo_mercadoria_trimestre = float(custo_mercadorias[f'MvtoLíquido Jan/{ano}'] + custo_mercadorias[f'MvtoLíquido Fev/{ano}'] + custo_mercadorias[f'MvtoLíquido Mar/{ano}'])
        custo_mercadoria_trimestre_porc = custo_mercadoria_trimestre / receita_trimestre

        #MARGEM BRUTA
        calculo_margem_bruta_janeiro = receita_operacional[f'MvtoLíquido Jan/{ano}'] + custo_mercadorias[f'MvtoLíquido Jan/{ano}']
        calculo_margem_bruta_fevereiro = receita_operacional[f'MvtoLíquido Fev/{ano}'] + custo_mercadorias[f'MvtoLíquido Fev/{ano}']
        calculo_margem_bruta_marco = receita_operacional[f'MvtoLíquido Mar/{ano}'] + custo_mercadorias[f'MvtoLíquido Mar/{ano}']
        margem_bruta_trimestre = calculo_margem_bruta_janeiro +  calculo_margem_bruta_fevereiro + calculo_margem_bruta_marco
        margem_bruta_trimestre_porc = margem_bruta_trimestre / receita_trimestre

        #Receitas Comerciais
        receitas_comerciais_trimestre = receitas_comercial[f'MvtoLíquido Jan/{ano}'] + receitas_comercial[f'MvtoLíquido Fev/{ano}'] + receitas_comercial[f'MvtoLíquido Mar/{ano}']
        receitas_comerciais_trimestre_porc = receitas_comerciais_trimestre / receita_trimestre

        #MARGEM COMERCIAL
        calculo_margem_janeiro = calculo_margem_bruta_janeiro + receitas_comercial[f'MvtoLíquido Jan/{ano}']
        calculo_margem_fevereiro = calculo_margem_bruta_fevereiro + receitas_comercial[f'MvtoLíquido Fev/{ano}']
        calculo_margem_marco = calculo_margem_bruta_marco + receitas_comercial[f'MvtoLíquido Mar/{ano}']
        calculo_margem_trimestre = calculo_margem_janeiro + calculo_margem_fevereiro + calculo_margem_marco
        calculo_margem_trimestre_porc = calculo_margem_trimestre / receita_trimestre

        #Quebra Contábil
        try:
            calculo_quebra_cotabil_janeiro = quebra_estoque[f'MvtoLíquido Jan/{ano}'] + ajuste_inventario[f'MvtoLíquido Jan/{ano}']
            calculo_quebra_cotabil_fevereiro = quebra_estoque[f'MvtoLíquido Fev/{ano}'] + ajuste_inventario[f'MvtoLíquido Fev/{ano}']
            calculo_quebra_cotabil_marco = quebra_estoque[f'MvtoLíquido Mar/{ano}'] + ajuste_inventario[f'MvtoLíquido Mar/{ano}']
            calculo_quebra_cotabil_trimestre = calculo_quebra_cotabil_janeiro + calculo_quebra_cotabil_fevereiro + calculo_quebra_cotabil_marco
        except:
            calculo_quebra_cotabil_janeiro = quebra_estoque[f'MvtoLíquido Jan/{ano}'] + perdas_mercadorias[f'MvtoLíquido Jan/{ano}']
            calculo_quebra_cotabil_Fevereiro = quebra_estoque[f'MvtoLíquido Fev/{ano}'] + perdas_mercadorias[f'MvtoLíquido Fev/{ano}']
            calculo_quebra_cotabil_marco = quebra_estoque[f'MvtoLíquido Mar/{ano}'] + perdas_mercadorias[f'MvtoLíquido Mar/{ano}']
            calculo_quebra_cotabil_trimestre = calculo_quebra_cotabil_janeiro + calculo_quebra_cotabil_Fevereiro + calculo_quebra_cotabil_marco
        calculo_quebra_cotabil_porc = calculo_quebra_cotabil_trimestre / receita_trimestre
        
        #

       

        primeiro_trimestre['RECEITA OPERACIONAL LÍQUIDA'] = receita_trimestre
        primeiro_trimestre['Custo de Mercadoria Vendida'] = custo_mercadoria_trimestre
        primeiro_trimestre['MARGEM BRUTA'] = margem_bruta_trimestre
        primeiro_trimestre['Receitas Comerciais'] = receitas_comerciais_trimestre
        primeiro_trimestre['MARGEM COMERCIAL'] = calculo_margem_trimestre
        primeiro_trimestre['Quebra Contábil'] = calculo_quebra_cotabil_trimestre
        primeiro_trimestre['Despesa C/ Venda'] = 0
        primeiro_trimestre['MARGEM OPERACIONAL'] = 0
        primeiro_trimestre['Despesas Operacionais'] = 0
        primeiro_trimestre['EBITDA'] = 0
        primeiro_trimestre['Depreciação'] = 0
        primeiro_trimestre['Juros Financeiros'] = 0
        primeiro_trimestre['Receita Financeiras'] = 0
        primeiro_trimestre['LAIR'] = 0
        primeiro_trimestre['IR'] = 0
        primeiro_trimestre['CSLL'] = 0
        primeiro_trimestre['LL - Lucro Líquido'] = 0

        primeiro_trimestre_porc['RECEITA OPERACIONAL LÍQUIDA'] = receita_trimestre_porc
        primeiro_trimestre_porc['Custo de Mercadoria Vendida'] = custo_mercadoria_trimestre_porc
        primeiro_trimestre_porc['MARGEM BRUTA'] = margem_bruta_trimestre_porc
        primeiro_trimestre_porc['Receitas Comerciais'] = receitas_comerciais_trimestre_porc
        primeiro_trimestre_porc['MARGEM COMERCIAL'] = calculo_margem_trimestre_porc
        primeiro_trimestre_porc['Quebra Contábil'] = calculo_quebra_cotabil_porc
        primeiro_trimestre_porc['Despesa C/ Venda'] = 0
        primeiro_trimestre_porc['MARGEM OPERACIONAL'] = 0
        primeiro_trimestre_porc['Despesas Operacionais'] = 0
        primeiro_trimestre_porc['EBITDA'] = 0
        primeiro_trimestre_porc['Depreciação'] = 0
        primeiro_trimestre_porc['Juros Financeiros'] = 0
        primeiro_trimestre_porc['Receita Financeiras'] = 0
        primeiro_trimestre_porc['LAIR'] = 0
        primeiro_trimestre_porc['IR'] = 0
        primeiro_trimestre_porc['CSLL'] = 0
        primeiro_trimestre_porc['LL - Lucro Líquido'] = 0

        #Trimestre
        preencher_valor_planilha(primeiro_trimestre)
        preencher_valor_planilha(primeiro_trimestre_porc)

    elif i == 2:
        preencher_dicionarios(segundo_trimestre, segundo_trimestre_porc, coluna_a_encontrar_tri, coluna_a_encontrar_tri_porc, False)
        #RECEITA OPERACIONAL LÍQUIDA
        receita_trimestre = float(receita_operacional[f'MvtoLíquido Abr/{ano}'] + receita_operacional[f'MvtoLíquido Mai/{ano}'] + receita_operacional[f'MvtoLíquido Jun/{ano}'])
        receita_trimestre_porc = receita_trimestre / receita_trimestre
        #Custo de Mercadoria Vendida
        custo_mercadoria_trimestre = float(custo_mercadorias[f'MvtoLíquido Abr/{ano}'] + custo_mercadorias[f'MvtoLíquido Mai/{ano}'] + custo_mercadorias[f'MvtoLíquido Jun/{ano}'])
        custo_mercadoria_trimestre_porc = custo_mercadoria_trimestre / receita_trimestre
        #MARGEM BRUTA
        calculo_margem_bruta_abril = receita_operacional[f'MvtoLíquido Abr/{ano}'] + custo_mercadorias[f'MvtoLíquido Abr/{ano}']
        calculo_margem_bruta_maio = receita_operacional[f'MvtoLíquido Mai/{ano}'] + custo_mercadorias[f'MvtoLíquido Mai/{ano}']
        calculo_margem_bruta_junho = receita_operacional[f'MvtoLíquido Jun/{ano}'] + custo_mercadorias[f'MvtoLíquido Jun/{ano}']
        margem_bruta_trimestre = calculo_margem_bruta_abril +  calculo_margem_bruta_maio + calculo_margem_bruta_junho
        margem_bruta_trimestre_porc = margem_bruta_trimestre / receita_trimestre
        #Receitas Comerciais
        receitas_comerciais_trimestre = receitas_comercial[f'MvtoLíquido Abr/{ano}'] + receitas_comercial[f'MvtoLíquido Mai/{ano}'] + receitas_comercial[f'MvtoLíquido Jun/{ano}']
        receitas_comerciais_trimestre_porc = receitas_comerciais_trimestre / receita_trimestre
        #MARGEM COMERCIAL
        calculo_margem_abril = calculo_margem_bruta_abril + receitas_comercial[f'MvtoLíquido Abr/{ano}']
        calculo_margem_maio = calculo_margem_bruta_maio + receitas_comercial[f'MvtoLíquido Mai/{ano}']
        calculo_margem_junho = calculo_margem_bruta_junho + receitas_comercial[f'MvtoLíquido Jun/{ano}']
        calculo_margem_trimestre = calculo_margem_abril + calculo_margem_maio + calculo_margem_junho
        calculo_margem_trimestre_porc = calculo_margem_trimestre / receita_trimestre

        segundo_trimestre['RECEITA OPERACIONAL LÍQUIDA'] = receita_trimestre
        segundo_trimestre['Custo de Mercadoria Vendida'] = custo_mercadoria_trimestre
        segundo_trimestre['MARGEM BRUTA'] = margem_bruta_trimestre
        segundo_trimestre['Receitas Comerciais'] = receitas_comerciais_trimestre
        segundo_trimestre['MARGEM COMERCIAL'] = calculo_margem_trimestre
        segundo_trimestre['Quebra Contábil'] = 0
        segundo_trimestre['Despesa C/ Venda'] = 0
        segundo_trimestre['MARGEM OPERACIONAL'] = 0
        segundo_trimestre['Despesas Operacionais'] = 0
        segundo_trimestre['EBITDA'] = 0
        segundo_trimestre['Depreciação'] = 0
        segundo_trimestre['Juros Financeiros'] = 0
        segundo_trimestre['Receita Financeiras'] = 0
        segundo_trimestre['LAIR'] = 0
        segundo_trimestre['IR'] = 0
        segundo_trimestre['CSLL'] = 0
        segundo_trimestre['LL - Lucro Líquido'] = 0

        segundo_trimestre_porc['RECEITA OPERACIONAL LÍQUIDA'] = receita_trimestre_porc
        segundo_trimestre_porc['Custo de Mercadoria Vendida'] = custo_mercadoria_trimestre_porc
        segundo_trimestre_porc['MARGEM BRUTA'] = margem_bruta_trimestre_porc
        segundo_trimestre_porc['Receitas Comerciais'] = receitas_comerciais_trimestre_porc
        segundo_trimestre_porc['MARGEM COMERCIAL'] = calculo_margem_trimestre_porc
        segundo_trimestre_porc['Quebra Contábil'] = 0
        segundo_trimestre_porc['Despesa C/ Venda'] = 0
        segundo_trimestre_porc['MARGEM OPERACIONAL'] = 0
        segundo_trimestre_porc['Despesas Operacionais'] = 0
        segundo_trimestre_porc['EBITDA'] = 0
        segundo_trimestre_porc['Depreciação'] = 0
        segundo_trimestre_porc['Juros Financeiros'] = 0
        segundo_trimestre_porc['Receita Financeiras'] = 0
        segundo_trimestre_porc['LAIR'] = 0
        segundo_trimestre_porc['IR'] = 0
        segundo_trimestre_porc['CSLL'] = 0
        segundo_trimestre_porc['LL - Lucro Líquido'] = 0

        #Trimestre
        preencher_valor_planilha(segundo_trimestre)
        preencher_valor_planilha(segundo_trimestre_porc)
    elif i == 3:
        # Chamadas de função para preencher os dicionários
        preencher_dicionarios(terceiro_trimestre, terceiro_trimestre_porc, coluna_a_encontrar_tri, coluna_a_encontrar_tri_porc, False)

        #Trimestre
        preencher_valor_planilha(terceiro_trimestre)
    elif i == 4:
        # Chamadas de função para preencher os dicionários
        preencher_dicionarios(quarto_trimestre, quarto_trimestre_porc, coluna_a_encontrar_tri, coluna_a_encontrar_tri_porc, False)

        #Trimestre
        preencher_valor_planilha(quarto_trimestre)

# Salvar planilhas Excel tratadas, trocando index = True mostra o index das linhas
with pd.ExcelWriter(f'dre_tratada.xlsx') as writer:
    df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
    df_sheet1.to_excel(writer, sheet_name='Planilha DRE', index=False)

# Finalizar o temporizador
end_time = time.time()
# Calcular o tempo de execução
elapsed_time = end_time - start_time

print(f"Tempo total de execução: {elapsed_time:.2f} segundos")

# Carregar o arquivo Excel tratado
wb = load_workbook('dre_tratada.xlsx')
# Especifique o nome da aba que deseja modificar
nome_da_aba = 'Planilha DRE'
# Acesse a aba desejada
aba_existente = wb[nome_da_aba]

def cor_fundo_celula(numero_index_linha, cor_hex):
    # Especifique a cor de fundo que deseja aplicar #ffffff
    cor_de_fundo = cor_hex 
    # Percorra todas as células da linha especificada e defina a cor de fundo
    for cell in aba_existente[numero_index_linha]:
        cell.fill = PatternFill(start_color=cor_de_fundo, end_color=cor_de_fundo, fill_type='solid')
    
def encontrar_indices_texto(linha_descricao):
# Preencher os valores ausentes na coluna de interesse com um valor padrão
    df_filled = df.fillna({'DescricaoConta ': ''})
    # Localizar as células que contêm o texto inserido
    linhas_contas_resultado = df_filled[df_filled['DescricaoConta '].str.contains(pat=linha_descricao)]
    return linhas_contas_resultado.index.tolist()

lista_descrição_azul = (
'CONTAS DE RESULTADO', 'RECEITAS', 'RECEITAS OPERACIONAL LIQUIDA', 'DEDUCOES DA VENDA DE MERCADORIAS','VENDA DE MERCADORIAS' ,
'RECEITAS COMERCIAIS', 'RECEITAS DIVERSAS', 'RECEITAS FINANCEIRAS', 'RECEITAS PRESTACOES SERVICOS',
'CUSTO DOS SERVICOS PRESTADOS', 'CUSTO DOS SERVIÇOS PRESTADOS' , 'CUSTOS DAS MERCADORIAS VENDIDOS', 'CUSTO DAS MERCADORIAS VENDIDOS - CMV',
'ESTOQUE INICIAL', 'ENTRADAS DE MERCADORIAS', 'ESTOQUE FINAL', 'CUSTO OPERACIONAL', 'CUSTO DE PRODUCAO',
'DESPESAS OPERACIONAIS', 'DESPESAS COM PESSOAL', 'PRO LABORE', 'DESPESAS COM PESSOAL DIRETO', 'BENEFICIOS',
'REMUNERACAO VARIAVEIS', 'SERVICOS TERCEIRIZADOS', 'OUTRAS DESPESAS OPERACIONAIS', 'MANUTENCAO', 'SEGUROS',
'DESPESAS COM VEICULOS', 'TRANSPORTES E DESLOCAMENTOS', 'VIAGENS E ESTADIAS', 'ENERGIA', 'COMUNICACAO',
'OUTRAS DESPESAS VENDAS', 'DESPESAS EMBALAGENS', 'DESPESAS CONSUMO INTERNO', 'MANUTENCAO DE IMOVEIS',
'DESPESAS COM INFORMATICA', 'DESPESAS GERAIS', 'MARKETING', 'DESPESAS DE EXPEDIENTE', 'ALUGUEIS',
'IMPOSTOS E TAXAS', 'DEPRECIACOES E PROVISOES', 'DESPESAS DIVERSAS', 'HONORARIOS PROFISSIONAIS',
'DESPESAS FINANCEIRAS', 'SUPERMERCADO AQUINO', 'ASSOCIACOES DE CLASSE', 'DESPESAS FISCAIS',
'RESULTADO DO EXERCICIO', 'DESPESAS E RECEITAS', 'RECEITAS OPERACIONAIS', 'DESPESAS OPERACIONAIS',
'CUSTO DA MERCADORIAS VENDIDAS', 'RECEITAS E DESPESAS NAO OPERACIONAIS', 'RESULTADO NA VENDA IMOBILIZADO',
'PROVISAO DE IMPOSTO S/L', 'PROVISAO DE IMPOSTO S/L', 'TRANSF. PARA RESERVA DE L', 'TRANFERENCIA PARA RESERVA DE L'
)

# faz um loop for com a lista para pintar as linhas cuja descrição tem na lista
for descricao in lista_descrição_azul:
    indices = encontrar_indices_texto(descricao)
    for index in indices:
        index_int = int(index)  # O deslocamento de 2 é para corresponder às linhas do DataFrame df_resumo
        cor_fundo_celula(index_int, 'C5D9F1')

cor_fundo_celula(4, 'FFFF00') #amarelo

# Salvar novamente o arquivo
wb.save('dre_tratada.xlsx')


    
   